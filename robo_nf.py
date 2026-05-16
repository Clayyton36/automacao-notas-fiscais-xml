#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import xmltodict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Caminhos adaptados para Linux
pasta_principal = os.path.abspath(os.path.dirname(__file__))
pasta_notas = os.path.join(pasta_principal, "notas")
arquivo_saida_base = os.path.join(pasta_principal, "Planilha_Notas_Fiscais")

dados = []

print("🔍 Iniciando leitura das notas...\n")

# Verificar se a pasta existe
if not os.path.isdir(pasta_notas):
    print("❌ ERRO: A pasta 'notas' não foi encontrada! Crie ela e coloque os XMLs dentro.")
    input("Aperte ENTER para sair...")
    exit()

# LER XMLS - VERSÃO COMPLETA, PEGA TUDO
for nome_arquivo in os.listdir(pasta_notas):
    if nome_arquivo.lower().endswith(".xml"):
        caminho = os.path.join(pasta_notas, nome_arquivo)
        try:
            # Leitura segura para qualquer codificação
            with open(caminho, "rb") as f:
                conteudo = f.read().decode("utf-8", errors="ignore")
            xml = xmltodict.parse(conteudo)

            # Aceita QUALQUER estrutura de XML
            nfe = None
            if "nfeProc" in xml:
                nfe = xml["nfeProc"].get("NFe", {}).get("infNFe", {})
            elif "NFe" in xml:
                nfe = xml["NFe"].get("infNFe", {})
            elif "infNFe" in xml:
                nfe = xml["infNFe"]

            if not nfe:
                print(f"⚠️ AVISO: {nome_arquivo} não tem estrutura de NF válida")
                continue

            # ==== DADOS BÁSICOS ====
            # Número da NF
            numero_nf = "NÃO ENCONTRADO"
            if "ide" in nfe:
                numero_nf = nfe["ide"].get("nNF", "NÃO ENCONTRADO")

            # Data de Emissão
            data_emis_texto = ""
            if "ide" in nfe:
                data_emis_texto = nfe["ide"].get("dhEmi", "")[:10] or nfe["ide"].get("dEmi", "")
            data_emis = ""
            if data_emis_texto:
                try:
                    data_emis = datetime.strptime(data_emis_texto, "%Y-%m-%d").strftime("%d/%m/%Y")
                except:
                    data_emis = "DATA INVÁLIDA"

            # Fornecedor
            fornecedor = "NÃO ENCONTRADO"
            if "emit" in nfe:
                nome_completo = nfe["emit"].get("xNome", "")
                fornecedor = nome_completo.split()[0] if nome_completo else "NÃO ENCONTRADO"

            # Valor Total
            valor_total = 0.0
            if "total" in nfe and "ICMSTot" in nfe["total"]:
                try:
                    valor_total = float(nfe["total"]["ICMSTot"].get("vNF", 0.0))
                except:
                    valor_total = 0.0

            # ==== PARCELAMENTO ====
            tem_parcela = False
            if "cobr" in nfe and nfe["cobr"] is not None and "dup" in nfe["cobr"]:
                parcelas = nfe["cobr"]["dup"]
                if isinstance(parcelas, dict):
                    parcelas = [parcelas]
                if isinstance(parcelas, list) and len(parcelas) > 0:
                    qtd = len(parcelas)
                    for i, p in enumerate(parcelas, 1):
                        if not p: continue
                        try:
                            v_parc = float(p.get("vDup", 0.0))
                        except:
                            v_parc = 0.0
                        dt_venc_texto = p.get("dVenc", "")
                        dt_venc = ""
                        if dt_venc_texto:
                            try:
                                dt_venc = datetime.strptime(dt_venc_texto, "%Y-%m-%d").strftime("%d/%m/%Y")
                            except:
                                dt_venc = "INVÁLIDO"
                        dias = "-"
                        if data_emis_texto and dt_venc_texto:
                            try:
                                dt_emi = datetime.strptime(data_emis_texto, "%Y-%m-%d")
                                dt_ven = datetime.strptime(dt_venc_texto, "%Y-%m-%d")
                                dias = (dt_ven - dt_emi).days
                            except:
                                dias = "-"
                        dados.append([numero_nf, data_emis, fornecedor, valor_total, qtd, i, v_parc, dt_venc, dias])
                    tem_parcela = True

            if not tem_parcela:
                dados.append([numero_nf, data_emis, fornecedor, valor_total, "Sem parcela", "-", "-", "-", "-"])

            print(f"✅ OK: {nome_arquivo}")

        except Exception as e:
            print(f"ℹ️ AVISO {nome_arquivo}: {str(e)}")

# CRIAR E SALVAR PLANILHA
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    # Cabeçalho
    cabecalhos = ["Número NF", "Data Emissão", "Fornecedor", "Valor Total", "Qtd Parcelas", "Nº Parcela", "Valor Parcela", "Vencimento", "Dias Pagto"]
    ws.append(cabecalhos)

    # Estilo bonito
    cor_cab = PatternFill("solid", fgColor="2F4F4F")
    fonte_cab = Font(color="FFFFFF", bold=True)
    alinhar = Alignment(horizontal="center", vertical="center")
    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for celula in ws[1]:
        celula.fill = cor_cab
        celula.font = fonte_cab
        celula.alignment = alinhar
        celula.border = borda

    # Dados
    for linha in dados:
        ws.append(linha)

    # Formatar valores em R$
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            celula.alignment = alinhar
            celula.border = borda
            if celula.column_letter in ("D", "G") and isinstance(celula.value, (int, float)) and celula.value > 0:
                celula.number_format = 'R$ #,##0.00'

    # Ajustar tamanho das colunas
    for col in ws.columns:
        largura = max(len(str(c.value)) for c in col if c.value is not None) + 3
        ws.column_dimensions[col[0].column_letter].width = largura

    # Salvar: evita erro se arquivo já existir
    arquivo_saida = arquivo_saida_base + ".xlsx"
    if os.path.exists(arquivo_saida):
        arquivo_saida = arquivo_saida_base + f"_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    wb.save(arquivo_saida)

    print(f"\n🎉 TUDO PRONTO!")
    print(f"📂 Arquivo salvo em:\n{arquivo_saida}")
    print(f"📊 Total de notas processadas: {len(dados)}")

except Exception as erro_final:
    print(f"\n❌ ERRO AO GERAR: {erro_final}")

input("\nAperte ENTER para fechar...")

